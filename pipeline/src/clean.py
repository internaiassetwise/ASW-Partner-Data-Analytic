"""
Phase 1 (clean) — unify every sheet into one tidy long-form table.

Logic (why this order):
  1. Read each sheet with its known header row (from profile).
  2. Map header -> unified field via a normalized-name lookup.
  3. Drop fully empty rows.
  4. Normalize values (trim whitespace, parse numbers/dates).
  5. Parse the projects file separately (it has lat/lng already).
  6. Assign entity_type, project_zone, source_file/sheet/row.
  7. Generate a stable external_id from content (for upsert in Phase 5).
  8. Write ONE combined CSV: data/clean/partners_clean.csv
  9. Write a summary report for Gate 1.

Does NOT touch source files. Re-runnable.
"""
import hashlib
import json
import re
from datetime import datetime, date

import openpyxl
import pandas as pd

from config import B2B_FILE, PROJECTS_FILE, CLEAN_DIR, REPORTS_DIR

STRIP_CHARS = re.compile(r"[\s\-\[\]()/.,:;+'\"!?]+")

UNIFIED_FIELDS = [
    "external_id", "entity_type", "name", "email", "phone", "contact_name",
    "street", "city", "address_full", "employee_count", "join_date",
    "project_zone", "admin_zone", "subzone", "province",
    "lat", "lng", "geo_source", "status", "line_id", "follow",
    "m_intranet", "m_edm", "m_line", "m_standee", "m_poster", "m_booth", "m_leaflet",
    "remark", "notes",
    "source_file", "source_sheet", "source_row", "raw_data",
]

HEADER_MAP = {
    "email": "email",
    "suppliername": "name", "bankname": "name", "partnersname": "name",
    "street": "street", "address": "street",
    "city": "city",
    "contactperson": "contact_name", "salesperson": "contact_name",
    "telephone": "phone",
    "จำนวนพนักงานheadoffice": "employee_count",
    "จำนวนบุคลากร": "employee_count",
    "remark": "remark",
    "วันที่เข้าร่วม": "join_date",
    "สื่อสารผ่านระบบintranet": "m_intranet",
    "companywideemailedm": "m_edm",
    "lineบริษัท": "m_line",
    "lineid": "line_id",
    "standee": "m_standee",
    "poster": "m_poster",
    "booth": "m_booth",
    "leafletใบปลิว": "m_leaflet",
    "ใบปลิว": "m_leaflet",
    "inzoneproject": "project_zone",
    "follow": "follow",
}

# Sheet-specific overrides (header row + entity_type). Default header_row=2.
SHEET_META = {
    ("B2B", "ASW Partner"):        {"entity_type": "partner", "header_row": 2},
    ("B2B", "Sponsper"):           {"entity_type": "sponsor", "header_row": 2},
    ("B2B", "Bank"):               {"entity_type": "bank", "header_row": 2},
    ("B2B", "องค์กร ภายนอก"):       {"entity_type": "external_org", "header_row": 2},
    ("B2B", "2026"):               {"entity_type": "partner_2026", "header_row": 2},
    ("B2B", "กลุ่มข้าราชการ กทม."):  {"entity_type": "gov_bkk", "header_row": 1},
    ("B2B", "สำนักงานเขต"):         {"entity_type": "gov_district", "header_row": 1},
}

# Rows that look like entities but are actually section labels / artifacts in
# the source spreadsheet. Dropped during clean. Add more as discovered.
JUNK_NAMES = {"88", "Operative"}

# Canonical province/city names — every variant maps to ONE official name so
# that filtering on the website yields a single bucket per province.
CITY_CANONICAL = {
    # Bangkok
    "กรุงเทพฯ": "กรุงเทพมหานคร",
    "กรุงเทพมหานคร": "กรุงเทพมหานคร",
    "กทม": "กรุงเทพมหานคร",
    "กทม.": "กรุงเทพมหานคร",
    "bangkok": "กรุงเทพมหานคร",
    # Pathum Thani (typos found in data)
    "ปทุมธานี": "ปทุมธานี",
    "ปทุมธานีี": "ปทุมธานี",
    "ปทุธานี": "ปทุมธานี",
    # Ayutthaya (short vs full)
    "อยุธยา": "พระนครศรีอยุธยา",
    "พระนครศรีอยุธยา": "พระนครศรีอยุธยา",
}


def normalize_city(raw):
    """Map messy city strings -> canonical Thai province name."""
    if raw is None:
        return ""
    s = re.sub(r"\s+", "", str(raw).strip())
    if not s:
        return ""
    low = s.lower()
    if low in CITY_CANONICAL:
        return CITY_CANONICAL[low]
    if s in CITY_CANONICAL:
        return CITY_CANONICAL[s]
    # strip leading 'จ.' or 'จ .' province prefix
    m = re.match(r"^จ\.?\s*(.+)$", s)
    if m:
        rest = m.group(1)
        return CITY_CANONICAL.get(rest, rest)
    return s


def hkey(s):
    if s is None:
        return ""
    return STRIP_CHARS.sub("", str(s)).lower()


def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return re.sub(r"\s+", " ", s)


def parse_int(v):
    s = norm(v)
    if not s:
        return ""
    m = re.search(r"-?\d[\d,]*", s)
    if not m:
        return ""
    try:
        return int(m.group(0).replace(",", ""))
    except ValueError:
        return ""


def parse_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = norm(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_latlng(v):
    """Parse a 'lat, lng' string -> (lat, lng) floats or ('', '')."""
    s = norm(v)
    if not s:
        return "", ""
    m = re.findall(r"-?\d+\.?\d*", s)
    if len(m) >= 2:
        try:
            return float(m[0]), float(m[1])
        except ValueError:
            return "", ""
    return "", ""


def make_external_id(row):
    base = "|".join([
        row.get("entity_type", ""),
        row.get("name", ""),
        row.get("street", ""),
        row.get("city", ""),
        row.get("phone", ""),
        row.get("email", ""),
    ])
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def empty_row():
    return {f: "" for f in UNIFIED_FIELDS}


def load_matrix(ws):
    """Read whole sheet once -> {row_number: tuple} + real_max_column.
    Avoids slow per-cell access (esp. sheets with 16k reported columns)."""
    matrix = {}
    real_max = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        matrix[i] = row
        for j in range(len(row) - 1, -1, -1):
            v = row[j]
            if v is not None and str(v).strip():
                if j + 1 > real_max:
                    real_max = j + 1
                break
    return matrix, real_max


def cell(matrix, r, c):
    row = matrix.get(r)
    if not row or c - 1 >= len(row):
        return None
    return row[c - 1]


def clean_b2b_sheet(ws, meta, source_file):
    matrix, real_cols = load_matrix(ws)
    header_row = meta["header_row"]
    entity_type = meta["entity_type"]
    # Build col index -> unified field
    col_map = {}
    for c in range(1, real_cols + 1):
        h = cell(matrix, header_row, c)
        k = hkey(h)
        if k in HEADER_MAP:
            col_map[c] = HEADER_MAP[k]

    # Unmapped columns that still hold data -> capture as notes
    note_cols = []
    for c in range(1, real_cols + 1):
        if c in col_map:
            continue
        has_data = any(
            (cell(matrix, r, c) not in (None, ""))
            and str(cell(matrix, r, c)).strip()
            for r in range(header_row + 1, len(matrix) + 1)
        )
        if has_data:
            note_cols.append(c)

    rows = []
    for r in range(header_row + 1, len(matrix) + 1):
        rec = empty_row()
        rec["entity_type"] = entity_type
        rec["source_file"] = source_file
        rec["source_sheet"] = ws.title
        rec["source_row"] = r
        raw = {}
        for c in range(1, real_cols + 1):
            v = cell(matrix, r, c)
            raw[str(c)] = "" if v is None else str(v)
            if c in col_map:
                field = col_map[c]
                if field == "employee_count":
                    rec[field] = parse_int(v)
                elif field == "join_date":
                    rec[field] = parse_date(v)
                elif field == "city":
                    rec[field] = normalize_city(norm(v))
                else:
                    rec[field] = norm(v)
        note_parts = [norm(cell(matrix, r, c)) for c in note_cols]
        note_parts = [p for p in note_parts if p]
        rec["notes"] = " ; ".join(note_parts)
        if not any(rec[k] for k in ("name", "email", "phone", "street", "contact_name")):
            continue
        if rec["name"] in JUNK_NAMES:
            continue
        rec["address_full"] = ", ".join(p for p in [rec["street"], rec["city"]] if p)
        rec["geo_source"] = ""
        rec["raw_data"] = json.dumps(raw, ensure_ascii=False)
        rec["external_id"] = make_external_id(rec)
        rows.append(rec)
    return rows


def clean_projects(ws, source_file):
    """Projects file: header row 1, data starts row 4. lat/lng already present."""
    matrix, _ = load_matrix(ws)
    rows = []
    for r in range(4, len(matrix) + 1):
        idx = cell(matrix, r, 1)
        name = norm(cell(matrix, r, 2))
        bu = norm(cell(matrix, r, 3))
        status = norm(cell(matrix, r, 4))
        gmap = norm(cell(matrix, r, 5))
        zone = norm(cell(matrix, r, 6))
        addr = norm(cell(matrix, r, 7))
        latlng = cell(matrix, r, 8)
        if not name and idx in (None, ""):
            continue
        lat, lng = parse_latlng(latlng)
        # Section headers in the projects file (zone labels like "รัชดา-ลาดพร้าว",
        # "EEC", "Housing", "Operative") have NO coordinates. Real projects all do.
        # Drop any row without a parseable lat/lng -> removes ~7 junk section rows.
        if lat == "" or lng == "":
            continue
        rec = empty_row()
        rec.update({
            "entity_type": "project",
            "name": name,
            "street": addr,
            "address_full": addr,
            "project_zone": zone,
            "status": status,
            "lat": lat,
            "lng": lng,
            "geo_source": "file" if (lat != "" and lng != "") else "",
            "remark": gmap,
            "notes": f"BU={bu}",
            "source_file": source_file,
            "source_sheet": ws.title,
            "source_row": r,
            "raw_data": json.dumps({
                "1": "" if idx is None else str(idx), "2": name, "3": bu, "4": status,
                "5": gmap, "6": zone, "7": addr, "8": "" if latlng is None else str(latlng),
            }, ensure_ascii=False),
        })
        rec["external_id"] = make_external_id(rec)
        rows.append(rec)
    return rows


def main():
    all_rows = []

    wb = openpyxl.load_workbook(B2B_FILE, data_only=True)
    for ws in wb.worksheets:
        key = ("B2B", ws.title)
        if key not in SHEET_META:
            print(f"  skip sheet: {ws.title!r} (not in SHEET_META)")
            continue
        meta = SHEET_META[key]
        n_before = ws.max_row - meta["header_row"]
        rows = clean_b2b_sheet(ws, meta, "B2B")
        all_rows.extend(rows)
        print(f"  {ws.title!r}: scanned {n_before} -> kept {len(rows)}")
    wb.close()

    wb = openpyxl.load_workbook(PROJECTS_FILE, data_only=True)
    ws = wb["Sheet1"]
    rows = clean_projects(ws, "PROJECTS")
    all_rows.extend(rows)
    print(f"  PROJECTS/Sheet1: kept {len(rows)} (lat/lng from file)")
    wb.close()

    df = pd.DataFrame(all_rows, columns=UNIFIED_FIELDS)
    out = CLEAN_DIR / "partners_clean.csv"
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"\nWrote {len(df)} rows -> {out}")

    # ---- Gate 1 summary report ----
    lines = ["# Phase 1 — Clean Summary (Gate 1)\n"]
    lines.append(f"Total rows: **{len(df)}**\n")
    lines.append("\n## Rows per entity_type\n")
    vc = df["entity_type"].value_counts()
    for k, v in vc.items():
        lines.append(f"- `{k}`: {v}")
    lines.append("\n## Rows per source_sheet\n")
    sc = df.groupby(["source_file", "source_sheet"]).size()
    for (f, s), v in sc.items():
        lines.append(f"- `{f}` / `{s}`: {v}")

    lines.append("\n## Geocode status before Phase 3\n")
    have_geo = df[(df["lat"] != "") & (df["lng"] != "")]
    need_geo = df[(df["lat"] == "") | (df["lng"] == "")]
    lines.append(f"- already have lat/lng (projects): **{len(have_geo)}**")
    lines.append(f"- need geocoding: **{len(need_geo)}**")

    lines.append("\n## project_zone fill rate per entity_type\n")
    pz = df.groupby("entity_type").apply(
        lambda g: f"{(g['project_zone'] != '').sum()}/{len(g)}"
    )
    for k, v in pz.items():
        lines.append(f"- `{k}`: {v}")

    lines.append("\n## Address coverage (need for geocoding)\n")
    for et in df["entity_type"].unique():
        g = df[df["entity_type"] == et]
        with_addr = (g["address_full"] != "").sum()
        lines.append(f"- `{et}`: {with_addr}/{len(g)} have address")

    lines.append("\n## Duplicate external_id check\n")
    dups = df[df["external_id"].duplicated(keep=False)]
    if len(dups) == 0:
        lines.append("- no duplicates")
    else:
        lines.append(f"- **{len(dups)} rows share ids** — review")
        for _, r in dups.head(10).iterrows():
            lines.append(f"  - {r['external_id']} | {r['entity_type']} | {r['name'][:40]}")

    lines.append("\n## Sample rows (first 2 per entity_type)\n")
    for et in df["entity_type"].unique():
        lines.append(f"\n### {et}")
        for _, r in df[df["entity_type"] == et].head(2).iterrows():
            lines.append(
                f"- name=`{r['name'][:40]}` city=`{r['city']}` zone=`{r['project_zone']}` "
                f"lat=`{r['lat']}` lng=`{r['lng']}` phone=`{r['phone'][:20]}`"
            )

    (REPORTS_DIR / "clean_summary.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote summary -> {REPORTS_DIR / 'clean_summary.md'}")


if __name__ == "__main__":
    print("Cleaning all sheets ...")
    main()
