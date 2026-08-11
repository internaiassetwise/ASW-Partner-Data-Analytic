"""
Phase 1 — Profile every sheet of both source Excel files.

Output:
  - reports/profile.md   (human-readable, UTF-8)
  - reports/profile.json (machine-readable)

The script does NOT modify any source file. It only inspects.
"""
import json
import sys
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter

from config import B2B_FILE, PROJECTS_FILE, REPORTS_DIR


def open_workbook(path):
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


def count_non_empty(row):
    return sum(1 for c in row if c is not None and str(c).strip() != "")


def detect_header_row(ws, scan=6):
    """Pick the row (1-indexed) within the first `scan` rows that has the most
    non-empty cells AND looks like text headers (not numbers)."""
    best_row, best_score = 1, -1
    for r in range(1, min(scan, ws.max_row) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = count_non_empty(row)
        text_like = sum(
            1
            for c in row
            if isinstance(c, str) and c.strip() and not str(c).strip().replace(".", "").replace("-", "").isdigit()
        )
        score = non_empty + (0.5 if text_like >= max(1, non_empty - 1) else 0)
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def real_max_column(ws):
    """Find the rightmost column that has any data across all rows."""
    last = 0
    for row in ws.iter_rows(values_only=True):
        for idx in range(len(row) - 1, -1, -1):
            if row[idx] is not None and str(row[idx]).strip() != "":
                if idx + 1 > last:
                    last = idx + 1
                break
    return last


def profile_sheet(ws):
    real_cols = real_max_column(ws)
    header_row = detect_header_row(ws)
    headers = [
        (ws.cell(row=header_row, column=c).value or f"col_{c}")
        for c in range(1, real_cols + 1)
    ]

    data_rows = []
    non_empty_count = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, real_cols + 1)]
        if count_non_empty(row) > 0:
            non_empty_count += 1
            if len(data_rows) < 3:
                data_rows.append(row)

    null_counts = [0] * real_cols
    total_data = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, real_cols + 1)]
        if count_non_empty(row) == 0:
            continue
        total_data += 1
        for i, v in enumerate(row):
            if v is None or str(v).strip() == "":
                null_counts[i] += 1

    null_pct = [
        round(100.0 * n / total_data, 1) if total_data else 0.0
        for n in null_counts
    ]

    col_non_empty = [
        total_data - null_counts[i] for i in range(real_cols)
    ]
    return {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_column_reported": ws.max_column,
        "real_max_column": real_cols,
        "header_row": header_row,
        "headers": headers,
        "data_rows_total": ws.max_row - header_row,
        "data_rows_non_empty": non_empty_count,
        "null_pct_per_col": null_pct,
        "col_non_empty": col_non_empty,
        "sample_rows": [
            [(str(c)[:40] if c is not None else "") for c in row] for row in data_rows
        ],
    }


def fmt_table(headers, rows, widths=None):
    if not rows:
        return "_(no rows)_\n"
    if widths is None:
        widths = []
        for i in range(len(headers)):
            w = len(str(headers[i]))
            for r in rows:
                if i < len(r):
                    w = max(w, min(40, len(str(r[i]))))
            widths.append(min(w + 2, 42))
    head = "".join(str(headers[i]).ljust(widths[i]) for i in range(len(headers)))
    sep = "".join("-" * widths[i] for i in range(len(headers)))
    body = []
    for r in rows:
        body.append("".join(str(r[i]).ljust(widths[i]) if i < len(r) else "".ljust(widths[i]) for i in range(len(headers))))
    return head + "\n" + sep + "\n" + "\n".join(body) + "\n"


def build_report(files):
    md = []
    json_out = {"files": []}

    md.append("# Phase 1 — Data Profile\n")
    for label, path in files:
        md.append(f"\n## File: `{path.name}`\n")
        wb = open_workbook(path)
        file_entry = {"file": str(path), "sheets": []}
        for ws in wb.worksheets:
            info = profile_sheet(ws)
            file_entry["sheets"].append(info)
            md.append(f"\n### Sheet: `{info['sheet_name']}`\n")
            md.append(
                f"- max_row (reported): **{info['max_row']}**, "
                f"real columns with data: **{info['real_max_column']}/{info['max_column_reported']}**\n"
                f"- header detected at row **{info['header_row']}**\n"
                f"- data rows: **{info['data_rows_total']}** total, "
                f"**{info['data_rows_non_empty']}** non-empty\n"
            )

            md.append("\n**Headers:**\n")
            md.append(" | ".join(f"`{h}`" for h in info["headers"]) + "\n")

            md.append("\n**Column fill rate** (non-empty / total):\n")
            fill_rows = [
                [info["headers"][i], info["col_non_empty"][i], info["null_pct_per_col"][i]]
                for i in range(len(info["headers"]))
            ]
            md.append("```\n")
            md.append(fmt_table(["column", "filled", "null%"], fill_rows))
            md.append("```\n")

            md.append("\n**Sample (first 3 data rows, truncated):**\n")
            md.append("```\n")
            md.append(fmt_table(info["headers"], info["sample_rows"]))
            md.append("```\n")
        json_out["files"].append(file_entry)
        wb.close()

    (REPORTS_DIR / "profile.md").write_text("".join(md), encoding="utf-8")
    (REPORTS_DIR / "profile.json").write_text(
        json.dumps(json_out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return md


if __name__ == "__main__":
    files = [("B2B", B2B_FILE), ("PROJECTS", PROJECTS_FILE)]
    print(f"Profiling {len(files)} files ...", flush=True)
    build_report(files)
    out = REPORTS_DIR / "profile.md"
    print(f"DONE -> {out}", flush=True)
